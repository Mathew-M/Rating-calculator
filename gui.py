from tkinter import *
from tkinter import ttk
from tkinter import messagebox
from tkinter import simpledialog
import openpyxl
import pandas as pd
import datetime
import dropbox
from dropbox.files import WriteMode
from ping3 import ping
from pathlib import Path
from openpyxl import Workbook


	
participants = "participants.csv"
before_participants = "reserve/before_participants.csv"
all_record = "all_record.csv"
before_record = "reserve/before_record.csv"

dropbox_access_token=""
dropbox_path_1 = '/rating/participants.csv'
dropbox_path_2 = '/rating/all_record.csv'

updown_pass = ''

def make_default():
	df_default_participants = pd.DataFrame(columns=['氏名', 'レート', '勝', '負', '勝率'])
	df_default_record = pd.DataFrame(columns=['日付', '勝者', '勝者対戦前レート', '勝者対戦後レート', '敗者', '敗者対戦前レート', '敗者対戦後レート', '変動', '時間残し'])
	try:
		df_default_participants.to_csv(participants, index=False)
		df_default_record.to_csv(all_record, index=False)
		df_default_participants.to_csv(before_participants, index=False)
		df_default_record.to_csv(before_record, index=False)
		xlsx_sync()
	except PermissionError:
		messagebox.showerror('権限エラー', '開かれている*.csvファイル、*.xlsxファイルを閉じてください。')


def rate_calc(result):
	df_participants = pd.read_csv(participants)
	try:
		df_participants.to_csv(before_participants, index=False) # undo用の保存
	except PermissionError:
		messagebox.showerror('権限エラー', '開かれているreserve/before_participants.csvを閉じてください。')

	df_result = result
	game_date = datetime.date.today()

	df_all_record = pd.read_csv(all_record)
	try:
		df_all_record.to_csv(before_record, index=False) # undo用の保存
	except PermissionError:
		messagebox.showerror('権限エラー', '開かれているreserve/before_record.csvを閉じてください。')

	name_list = df_participants["氏名"].to_list()
	for row in df_result.itertuples(name=None):
		winner_name, loser_name, time_rest = row[1], row[2], row[3]
		if winner_name in name_list and loser_name in name_list:
			winner_rate = df_participants.query('氏名 == @winner_name').iat[0, 1]
			loser_rate = df_participants.query('氏名 == @loser_name').iat[0, 1]

			if str(row[3]) == '1':
				fluctuation = round(((loser_rate-winner_rate+400)/25))*2
				time_rest = '1'
			else:
				fluctuation = round((loser_rate-winner_rate+400)/25)
				time_rest = '-'

			winner_new_rate = winner_rate + fluctuation
			loser_new_rate = loser_rate - fluctuation
			df_participants.loc[df_participants['氏名']==winner_name, 'レート'] = winner_new_rate
			df_participants.loc[df_participants['氏名']==loser_name, 'レート'] = loser_new_rate

			df_all_record.loc[len(df_all_record)] = [game_date, winner_name, winner_rate, winner_new_rate, loser_name, loser_rate, loser_new_rate, fluctuation, time_rest]
		else:
			pass

	
	for name in name_list:
		win = len(df_all_record.query("勝者 == @name"))
		lose = len(df_all_record.query("敗者 == @name"))
		df_participants.loc[df_participants['氏名']==name, '勝'] = win
		df_participants.loc[df_participants['氏名']==name, '負'] = lose
		if win+lose > 0:
			df_participants.loc[df_participants['氏名']==name, '勝率'] = win/(win+lose)
	try:
		df_participants.to_csv(participants, index=False)
		df_all_record.to_csv(all_record, index=False)
	except PermissionError:
		messagebox.showerror('権限エラー', '開かれているparticipants.csvまたはall_record.csvを閉じてください。')


root = Tk()
root.title('Rating Calculator v1.3')
root.geometry('910x500')


def reload():
	participants_show('')
	all_record_show('')
	filter_name()

def sort_participants(value):
	df_participants = pd.read_csv(participants)
	sorted_df = df_participants.sort_values(value, ascending=False)
	participants_show(sorted_df)

def participants_show(sorted_df):
	
	frame1 = ttk.Frame(root, width=500, padding=15)
	frame1.grid(row=0, column=0)

	part_label = ttk.Label(frame1, text='参加者一覧')
	part_label.grid(row=0, column=0)

	tree = ttk.Treeview(frame1, height=20)
	tree["columns"] = (0, 1, 2, 3, 4, 5)
	tree["show"] = "headings"
	tree.heading(0, text="ID", command=lambda: participants_show(''))
	tree.heading(1, text="氏名", command=lambda: sort_participants("氏名"))
	tree.heading(2, text="レート", command=lambda: sort_participants("レート"))
	tree.heading(3, text="勝", command=lambda: sort_participants("勝"))
	tree.heading(4, text="負", command=lambda: sort_participants("負"))
	tree.heading(5, text="勝率", command=lambda: sort_participants("勝率")) #printの代わりにコマンドにソートして表をリロードする関数を入れる
	tree.column(0, width=20)
	tree.column(1, width=80)
	tree.column(2, width=40)
	tree.column(3, width=30)
	tree.column(4, width=30)
	tree.column(5, width=40)

	tree.grid(row=1, column=0)
	if type(sorted_df) == str:
		df_participants = pd.read_csv(participants)
	else:
		df_participants = sorted_df
	for row in df_participants.itertuples(name=None):
		tree.insert("", "end", values=row)


def sort_all_record(value):
	df_all_record = pd.read_csv(all_record)
	sorted_df = df_all_record.sort_values(value, ascending=False)
	all_record_show(sorted_df)
	
def all_record_show(sorted_df):
	frame3 = ttk.Frame(root, padding=15)
	frame3.grid(row=0, column=1)

	part_label = ttk.Label(frame3, text='対局一覧')
	part_label.grid(row=0, column=0)

	tree = ttk.Treeview(frame3, height=20)
	tree["columns"] = (0, 1, 2, 3, 4, 5, 6, 7, 8)
	tree["show"] = "headings"
	tree.heading(0, text="日付", command=lambda: sort_all_record("日付"))
	tree.heading(1, text="勝者")
	tree.heading(2, text="対戦前")
	tree.heading(3, text="対戦後")
	tree.heading(4, text="敗者")
	tree.heading(5, text="対戦前")
	tree.heading(6, text="対戦後")
	tree.heading(7, text="変動", command=lambda: sort_all_record("変動"))
	tree.heading(8, text="時間残し")
	tree.column(0, width=80)
	tree.column(1, width=60)
	tree.column(2, width=45)
	tree.column(3, width=45)
	tree.column(4, width=60)
	tree.column(5, width=45)
	tree.column(6, width=45)
	tree.column(7, width=35)
	tree.column(8, width=33)

	tree.grid(row=1, column=0)

	if type(sorted_df) == str:
		df_all_record = pd.read_csv(all_record)
	else:
		df_all_record = sorted_df
	for row in df_all_record.itertuples(name=None, index=None):
		tree.insert("", "end", values=row)

def personal_filter(name):
	df_all_record = pd.read_csv(all_record)
	df_personal = df_all_record.query('勝者 == @name or 敗者 == @name')
	all_record_show(df_personal)

def make_result_add():
	result_window = Toplevel()
	result_window.title('結果追加')
	result_window.geometry("300x300")

	frame6 = ttk.Frame(result_window)
	frame6.grid(row=0, column=0, padx=40, pady=20)

	items = ['勝者', '敗者', '時間残し']
	for i in range(len(items)):
		lable_item = ttk.Label(frame6, text=items[i])
		lable_item.grid(row=0, column=i)

	n = 10

	df_participants = pd.read_csv(participants)
	names = tuple(df_participants['氏名'].to_list())

	items1 = ["0"] * n
	for i in range(n):
		items1[i] = ttk.Combobox(frame6, state='readonly', width=10)
		items1[i]['values'] = names
		items1[i].grid(row=i+1, column=0)

	items2 = ["0"] * n
	for i in range(n):
		items2[i] = ttk.Combobox(frame6, state='readonly', width=10)
		items2[i]['values'] = names
		items2[i].grid(row=i+1, column=1)

	items3 = ["0"] * n
	for i in range(n):
		items3[i] = StringVar()
		item3 = ttk.Checkbutton(frame6, onvalue="1", offvalue="0", variable=items3[i], width=1, padding=2)
		item3.grid(row=i+1, column=2, sticky=E)



	def result_entry():
		wb = openpyxl.load_workbook("result_add.xlsx")
		ws = wb['Result']

		result_data = []
		df_result2 = pd.DataFrame(columns=['winner', 'loser', 'time'])
		for i in range(n):
			result_data.append([items1[i].get(), items2[i].get(), items3[i].get()])
		for rec in result_data:
			if rec[0] != '' and rec[1] != '' and rec[0] != rec[1]:
				df_result2.loc[len(df_result2)] = [rec[0], rec[1], rec[2]]  # all_record.csv用
				ws.cell(row=ws.max_row+1, column=1).value = rec[0]  # cellのrowとcolumnのindexは1始まり
				ws.cell(row=ws.max_row, column=2).value = rec[1]  
				ws.cell(row=ws.max_row-1, column=3).value = rec[2]  
		try:
			wb.save("result_add.xlsx")
			rate_calc(df_result2)
		except PermissionError:
				messagebox.showerror('権限エラー', '開かれているresult_add.xlsxを閉じてください。')

		# df_result2.to_csv('result2.csv', index=False)
		print(df_result2)
		reload()
		result_window.destroy()

	def result_file_entry():
		df_participants = pd.read_csv(participants)
		df_all_record = pd.read_csv(all_record)
		df_result = pd.read_excel('result_add.xlsx')

		sharp_2 = df_result.query('winner == "##" | loser == "#"')
		if len(sharp_2) == 0:
			messagebox.showerror('入力エラー', '##が入力されていません')
		elif len(sharp_2) >= 2:
			messagebox.showerror('入力エラー', '##が複数行に入力されています')
		elif len(sharp_2) == 1:
			thre = int(df_result.query('winner == "##"').index[0])

			if thre == len(df_all_record):
				print("normal")
			if thre > len(df_all_record):
				print("error")
			if thre < len(df_all_record):
				print("overwrite")
				name_list = df_participants["氏名"].to_list()
				for name in name_list:
					for s in range(thre, len(df_all_record)):
						if df_all_record.at[s, "勝者"] == name:
							df_participants.loc[df_participants['氏名']==name, 'レート'] = df_all_record.at[s, "勝者対戦前レート"]
							break
						if df_all_record.at[s, "敗者"] == name:
							df_participants.loc[df_participants['氏名']==name, 'レート'] = df_all_record.at[s, "敗者対戦前レート"]
							break
						else:
							pass
			
			df_result = df_result[thre+1:]
			df_all_record = df_all_record[:thre]

			try:
				df_participants.to_csv(participants, index=False)
				df_all_record.to_csv(all_record, index=False)
				rate_calc(df_result)
			except PermissionError:
				messagebox.showerror('権限エラー', '開かれているparticipants.csvまたはall_record.csvを閉じてください。')

			print(df_result)

		reload()
		result_window.destroy()

	result_file_button = ttk.Button(frame6, text='ファイルで追加', padding=5, command=result_file_entry)
	result_file_button.grid(row=n+1, column=0)
	result_button = ttk.Button(frame6, text='直接追加', padding=5, command=result_entry)
	result_button.grid(row=n+1, column=1)

def make_user_add():
	user_window = Toplevel()
	user_window.title('ユーザー追加')
	user_window.geometry("250x300")

	frame7 = ttk.Frame(user_window)
	frame7.grid(row=0, column=0, padx=50, pady=20)

	items = ['氏名', '初期レート']
	for i in range(len(items)):
		lable_item = ttk.Label(frame7, text=items[i])
		lable_item.grid(row=0, column=i)

	n = 10

	df_participants = pd.read_csv(participants)
	names = tuple(df_participants['氏名'].to_list())

	items1 = ["0"] * n
	for i in range(n):
		items1[i] = StringVar()
		item1 = ttk.Entry(frame7, textvariable=items1[i], width=10)
		item1.grid(row=i+1, column=0)

	items2 = ["0"] * n
	for i in range(n):
		items2[i] = StringVar()
		item2 = ttk.Entry(frame7, textvariable=items2[i], width=10)
		item2.insert(0, "1500")
		item2.grid(row=i+1, column=1)


	def user_entry():
		df_participants =pd.read_csv(participants)
		try:
			df_participants.to_csv(before_participants, index=False) # undo用の保存
		except PermissionError:
			messagebox.showerror('権限エラー', '開かれているreserve/before_participants.csvを閉じてください。')
		df_all_record = pd.read_csv(all_record)
		try:
			df_all_record.to_csv(before_record, index=False) # undo用の保存
		except PermissionError:
			messagebox.showerror('権限エラー', '開かれているreserve/before_record.csvを閉じてください。')

		wb = openpyxl.load_workbook("result_add.xlsx")
		exist_user = df_participants['氏名'].to_list()
		user_add_data = []
		for i in range(n):
			user_add_data.append([items1[i].get(), items2[i].get()])
		valid = 0
		for user in user_add_data:
			for s in exist_user:
				if user[0] == s:
					valid += 1	
				else:
					pass
			if valid != 0:
				messagebox.showerror('入力エラー', '既に使われている名前があります。')
				break
			if user[0] != '':
				try:
					df_participants.loc[len(df_participants)] = [user[0], int(user[1]), 0, 0, 0]
					wb["Name_list"].cell(row=len(df_participants)+2, column=1).value = user[0]  # cellのrowとcolumnのindexは1始まり
				except ValueError:
					messagebox.showerror('入力エラー', '入力が不正な値です')
		try:
			wb.save("result_add.xlsx")
		except PermissionError:
			messagebox.showerror('権限エラー', '開かれているresult_add.xlsxを閉じてください。')

		try:
			df_participants.to_csv(participants, index=False)
		except PermissionError:
			messagebox.showerror('権限エラー', '開かれているparticipants.csvを閉じてください。')

		reload()
		user_window.destroy()

	result_button =ttk.Button(frame7, text='追加', padding=5, command=user_entry)
	result_button.grid(row=n+1, column=1)

def undo():
	df_before_participants = pd.read_csv(before_participants)
	df_before_record = pd.read_csv(before_record)
	df_participants = pd.read_csv(participants)
	df_all_record = pd.read_csv(all_record)
	df_before_participants.to_csv(participants, index=False)
	df_before_record.to_csv(all_record, index=False)
	df_participants.to_csv(before_participants, index=False)
	df_all_record.to_csv(before_record, index=False)
	reload()

def reset():
	confirm = messagebox.askyesno('確認', 'この操作は取り消せません。\n初期化しますか？')
	if confirm == True:
		make_default()
		reload()

def upload():
	ping_return = ping('yahoo.co.jp') # ネットワーク接続確認
	if ping_return == False:
		messagebox.showerror('エラー', 'インターネットに接続されていません。')
	else:
		upload_pass = simpledialog.askstring("アップロード", 'パスワードを入力してください')
		if upload_pass == updown_pass:
			client = dropbox.Dropbox(dropbox_access_token) # アカウントをリンク
			with open(participants, "rb") as f:
				client.files_upload(f.read(), dropbox_path_1, mode=WriteMode('overwrite'))
			with open(all_record, "rb") as f:
				client.files_upload(f.read(), dropbox_path_2, mode=WriteMode('overwrite'))
			messagebox.showinfo('確認', 'アップロード完了')
		elif upload_pass == None:
			pass
		else:
			messagebox.showerror('認証失敗', 'パスワードが間違っています')

def download():
	ping_return = ping('yahoo.co.jp')
	if ping_return == False:
		messagebox.showerror('エラー', 'インターネットに接続されていません。')
	else:
		download_pass = simpledialog.askstring("ダウンロード", 'パスワードを入力してください')
		if download_pass == updown_pass: 
			client = dropbox.Dropbox(dropbox_access_token) # アカウントをリンク
			with open(participants, "wb") as f:
				metadata, res = client.files_download(path=dropbox_path_1)
				f.write(res.content)
			with open(all_record, "wb") as f:
				metadata, res = client.files_download(path=dropbox_path_2)
				f.write(res.content)
			messagebox.showinfo('確認', 'ダウンロード完了')
			xlsx_sync()
			reload()
		elif download_pass == None:
			pass
		else:
			messagebox.showerror('認証失敗', 'パスワードが間違っています')

def bom_convert(path):
	with open(path, encoding='utf8') as f_in:
		data = f_in.read()
	try:
		with open(path, 'w', encoding='utf_8_sig') as f_out:
			f_out.write(data)
		messagebox.showinfo('確認', path+' BOM変換完了')
	except PermissionError:
		messagebox.showerror('権限エラー', '開かれている*.csvファイルを閉じてください。')

def nobom_convert(path):
	with open(path, encoding='utf_8_sig') as f_in:
		data = f_in.read()
	try:
		with open(path, 'w', encoding='utf8') as f_out:
			f_out.write(data)
	except PermissionError:
		messagebox.showerror('権限エラー', '開かれている*.csvファイルを閉じてください。')

def xlsx_sync():
	df_participants = pd.read_csv(participants)["氏名"].to_list()
	df_all_record = pd.read_csv(all_record)
	df_all_record_w = df_all_record["勝者"].to_list()
	df_all_record_l = df_all_record["敗者"].to_list()
	df_all_record_t = df_all_record["時間残し"].to_list()

	df_participants.insert(0, '##')

	df1 = pd.DataFrame(list(zip(df_all_record_w,df_all_record_l,df_all_record_t)), columns=['winner', 'loser', 'time'])
	df2 = pd.DataFrame(df_participants, columns=['氏名'])
	try:
		df1.to_excel('result_add.xlsx', sheet_name='Result', index=False)
		with pd.ExcelWriter('result_add.xlsx', mode='a') as writer:
			df2.to_excel(writer, sheet_name='Name_list', index=False)
		messagebox.showinfo('確認', 'xlsx同期完了')
	except PermissionError:
		messagebox.showerror('権限エラー', '開かれているresult_add.xlsxを閉じてください。')

def filter_name():
	frame4 = ttk.Frame(root)
	frame4.grid(row=0, column=2, sticky=N, pady=25)

	try:
		df_participants = pd.read_csv(participants)
	except:
		pass
	names = tuple(df_participants['氏名'].to_list())
	personal = ttk.Combobox(frame4, state='readonly', width=10)
	personal['values'] = names
	personal.grid(row=0, column=0)
	personal_button = ttk.Button(frame4, text='検索', padding=5, width=5, command=lambda: personal_filter(personal.get()))
	personal_button.grid(row=0, column=1)

	result_add_button = ttk.Button(frame4, text='結果追加', padding=5, width=12, command=make_result_add)
	result_add_button.grid(row=1, column=0, sticky=N)
	user_add_button = ttk.Button(frame4, text='ユーザー追加', padding=5, width=12, command=make_user_add)
	user_add_button.grid(row=2, column=0, sticky=N)
	undo_button = ttk.Button(frame4, text='直前操作取消', padding=5, width=12, command=undo)
	undo_button.grid(row=3, column=0, sticky=N)


	reload_button = ttk.Button(frame4, text='画面更新', padding=5, width=12, command=reload)
	reload_button.grid(row=4, column=0, sticky=N)

	reload_button = ttk.Button(frame4, text='UPLOAD', padding=5, width=12, command=upload)
	reload_button.grid(row=5, column=0, sticky=N)
	
	reload_button = ttk.Button(frame4, text='DOWNLOAD', padding=5, width=12, command=download)
	reload_button.grid(row=6, column=0, sticky=N)

	bom_convert_button = ttk.Button(frame4, text='BOM変換', padding=5, width=12, command=lambda:[bom_convert(participants), bom_convert(all_record)])
	bom_convert_button.grid(row=7, column=0, sticky=N)

	sync_button = ttk.Button(frame4, text='xlsx同期', padding=5, width=12, command=xlsx_sync)
	sync_button.grid(row=8, column=0, sticky=S)
	
	reset_button = ttk.Button(frame4, text='初期化', padding=5, width=12, command=reset)
	reset_button.grid(row=9, column=0, sticky=S)


if __name__ == "__main__":
	try:
		reload()
		root.mainloop()
	except FileNotFoundError:
		d = Path('reserve')
		d.mkdir()
		f = Path(before_participants)
		f.touch()
		f = Path(before_record)
		f.touch()
		make_default()
		reload()
		root.mainloop()
	except KeyError:
		nobom_convert(participants)
		nobom_convert(all_record)
		reload()
		root.mainloop()
